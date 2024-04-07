import openpyxl
import socket
import os

def dns_lookup(domain):
    try:
        ip = socket.gethostbyname(domain)
        return ip
    except socket.gaierror:
        return "N/A"

def process_excel(input_file):
    try:
        # Load the Excel file
        workbook = openpyxl.load_workbook(input_file)
        sheet = workbook.active

        # Start from the first row
        current_row = 1

        # Iterate over rows in the Excel file
        for row in sheet.iter_rows(min_row=1, max_col=1, max_row=sheet.max_row, values_only=True):
            domain = row[0]
            ip = dns_lookup(domain)
            # Write the IP address to the adjacent cell
            sheet.cell(row=current_row, column=2, value=ip)
            # Move to the next row
            current_row += 1

        # Save the changes
        output_file = os.path.splitext(input_file)[0] + '_with_ips.xlsx'
        workbook.save(output_file)
        print(f"IP addresses have been written to {output_file}")

    except Exception as e:
        print(f"An error occurred: {str(e)}")

if __name__ == "__main__":
    input_file = input("Enter the Excel file path (e.g., C:\\path\\to\\file.xlsx): ")
    process_excel(input_file)
